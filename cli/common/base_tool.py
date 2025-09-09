import os
import re
import openpyxl
import json
import xml.etree.ElementTree as ET
from cli.settings import is_debug
from cli.common.util.commit_utils import generate_cli_script_from_xml, load_param_dict, reverse_formula
from cli.settings import is_debug

class BaseTool():
    def __init__(self):
        super().__init__()

    def do_scf_to_cli(self, arg):
        """SCF XML을 CLI 명령어로 변환. 사용법: scf-to-cli <파일명>"""
        filename = arg.strip()
        if not filename:
            self.perror("사용법: scf-to-cli <xml파일명>")
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        xml_path = os.path.join(base_dir, "..", "data", "generated", filename)

        if not os.path.exists(xml_path):
            self.perror(f"[오류] 파일이 존재하지 않습니다: {xml_path}")
            return

        try:
            output_path = os.path.join(base_dir, "..", "data", "scripts", f"{os.path.splitext(filename)[0]}__script.cli")
            generate_cli_script_from_xml(xml_path, output_path)
            if is_debug:
                self.poutput(f"[완료] CLI 스크립트 저장됨: {output_path}")
        except Exception as e:
            self.perror(f"[오류] XML 파싱 중 문제 발생: {e}")

    ## 엔지니어
    def do_excel_to_dict_formula(self, arg):
        """
        엑셀 파일을 파싱해서 파라미터 변환 dict를 생성하고 저장합니다.
        사용법: excel_to_dict_formula <엑셀파일명>
        """
        excel_filename = arg.strip()
        if not excel_filename:
            self.perror("사용법: excel_to_dict_formula <엑셀파일명>")
            return

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        data_dir = os.path.join(base_dir, "data", "xlsx")
        excel_path = os.path.join(data_dir, excel_filename)

        if not os.path.exists(excel_path):
            self.perror(f"엑셀 파일이 존재하지 않습니다: {excel_path}")
            return

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb["Parameter List"]
        except KeyError:
            self.perror(f"엑셀 파일에 'Parameter List' 시트가 없습니다.")
            return
        except Exception as e:
            self.perror(f"엑셀 파일을 열 수 없습니다: {e}")
            return

        # 헤더 행 탐색: "Changes between releases"로 시작하는 행
        header_row_idx = None
        for row in sheet.iter_rows(min_row=1, max_row=20):
            cell_val = str(row[0].value).strip() if row[0].value else ""
            if cell_val.startswith("Changes between releases"):
                header_row_idx = row[0].row  # 0-based index
                print(f"[디버그] 헤더 행 탐색 완료 → 행 번호: {header_row_idx}")
                break

        if header_row_idx is None:
            self.perror("헤더 행을 찾을 수 없습니다.")
            return

        headers = sheet[header_row_idx]
        col_map = {}
        for idx, cell in enumerate(headers):
            val = str(cell.value).strip() if cell.value else ""
            if val == "Technology":
                col_map["tech"] = idx
            elif val == "Abbreviated Name":
                col_map["abbrev"] = idx
            elif val == "MO Class":
                col_map["mo"] = idx
            elif val == "Formula for Getting Internal Value":
                col_map["formula"] = idx
            elif val == "Default Value":
                col_map["default"] = idx

        required_keys = ["tech", "abbrev", "mo", "formula", "default"]
        missing_keys = [k for k in required_keys if k not in col_map]
        if missing_keys:
            self.perror(f"필수 헤더가 누락되었습니다: {missing_keys}")
            return

        param_dict = {}
        total_rows = sheet.max_row - header_row_idx

        for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1), start=1):
            tech = row[col_map["tech"]].value
            if tech not in ("SRAN", "4G", "5G", "LTE"):
                continue

            abbreviated_name = row[col_map["abbrev"]].value
            mo_class = row[col_map["mo"]].value
            formula = row[col_map["formula"]].value
            default_value = row[col_map["default"]].value

            if not abbreviated_name or not mo_class or not formula:
                continue

            if i % 100 == 0 or i == 1:
                percent = int((i / total_rows) * 100)
                if is_debug:
                    print(f"[{percent}%] {i}번째 파라미터 처리 중: {abbreviated_name}")

            mo_class = mo_class.strip()
            if '/' in mo_class:
                mo_class = mo_class.split('/')[-1]

            key = f"{mo_class}::{abbreviated_name.strip()}"
            param_dict[key] = {
                "formula": str(formula).strip(),
                "default": str(default_value).strip() if default_value is not None else None
            }

        version = "unknown"
        for part in excel_filename.split("_"):
            if part.startswith("24") or part.startswith("25"):
                version = part
                break

        output_filename = f"{version}_formula_param_dict.json"
        output_path = os.path.join(data_dir, output_filename)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(param_dict, f, indent=2, ensure_ascii=False)

        self.poutput(f"파라미터 변환 dict가 저장되었습니다: {output_path}")


    ## 엔지니어
    def do_excel_to_dict_mo(self, arg):
        """
        엑셀 파일을 파싱해서 MO 및 파라미터 유효성 검사용 dict를 생성하고 저장합니다.
        사용법: excel_to_dict_mo <엑셀파일명>
        """
        excel_filename = arg.strip()
        if not excel_filename:
            self.perror("사용법: excel_to_dict_mo <엑셀파일명>")
            return

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        data_dir = os.path.join(base_dir, "data", "xlsx")
        excel_path = os.path.join(data_dir, excel_filename)

        if not os.path.exists(excel_path):
            self.perror(f"엑셀 파일이 존재하지 않습니다: {excel_path}")
            return

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb["Parameter List"]
        except KeyError:
            self.perror(f"엑셀 파일에 'Parameter List' 시트가 없습니다.")
            return
        except Exception as e:
            self.perror(f"엑셀 파일을 열 수 없습니다: {e}")
            return

        # 헤더 행 탐색
        header_row_idx = None
        for row in sheet.iter_rows(min_row=1, max_row=20):
            cell_val = str(row[0].value).strip() if row[0].value else ""
            if cell_val.startswith("Changes between releases"):
                header_row_idx = row[0].row
                break
        if header_row_idx is None:
            self.perror("헤더 행을 찾을 수 없습니다.")
            return

        headers = sheet[header_row_idx]
        col_map = {}

        for idx, cell in enumerate(headers):
            val = str(cell.value).strip() if cell.value else ""
            if val == "Technology":
                col_map["tech"] = idx
            elif val == "Abbreviated Name":
                col_map["param_name"] = idx
            elif val == "MO Class":
                col_map["mo_path"] = idx
            elif val == "Parameter Category":
                col_map["param_category"] = idx
            elif val == "Parent Structure":
                col_map["parent_structure"] = idx
            elif val == "Data Type":
                col_map["param_type"] = idx
            elif val == "Multiplicity":
                col_map["multiplicity"] = idx
            elif val == "Range and step":
                col_map["range"] = idx
            elif val == "Default Value":
                col_map["default_value"] = idx

        required_keys = [
            "tech", "param_name", "mo_path", "param_category",
            "parent_structure", "param_type", "multiplicity", "range", "default_value"
        ]
        missing = [k for k in required_keys if k not in col_map]
        if missing:
            self.perror(f"필수 헤더가 누락되었습니다: {missing}")
            return

        mo_param_dict = {}
        raw_rows = []
        total_rows = sheet.max_row - header_row_idx

        for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1), start=1):
            if row[0].value == "Removed":
                continue

            tech = row[col_map["tech"]].value
            if tech not in ("SRAN", "4G", "5G", "LTE"):
                continue

            param_name = row[col_map["param_name"]].value
            mo_path = row[col_map["mo_path"]].value
            param_category = row[col_map["param_category"]].value
            parent_structure = row[col_map["parent_structure"]].value
            param_type = row[col_map["param_type"]].value
            multiplicity = row[col_map["multiplicity"]].value
            param_range = row[col_map["range"]].value
            default_value = row[col_map["default_value"]].value

            if not param_name or not mo_path:
                continue

            if i % 100 == 0 or i == 1:
                percent = int((i / total_rows) * 100)
                print(f"[{percent}%] {i}번째 파라미터 처리 중...")

            mo_parts = mo_path.strip().split("/")
            mo_class = mo_parts[-1]
            parent_mo = mo_parts[-2] if len(mo_parts) > 1 else None

            raw_rows.append({
                "mo_class": mo_class,
                "parent_mo": parent_mo,
                "param_name": param_name,
                "parent_structure": parent_structure.strip() if parent_structure else None,
                "param_type": param_type,
                "default_value": str(default_value).strip() if default_value else None,
                "required": (param_category == "Basic parameters"),
                "multiplicity": str(multiplicity).strip() if multiplicity else "1",
                "range": str(param_range).strip() if param_range else None
            })

        for row in raw_rows:
            mo = row["mo_class"]
            parent_mo = row["parent_mo"]
            pname = row["param_name"]
            parent = row["parent_structure"]

            if mo not in mo_param_dict:
                mo_param_dict[mo] = {"children": [], "params": {}}

            if parent_mo:
                if parent_mo not in mo_param_dict:
                    mo_param_dict[parent_mo] = {"children": [], "params": {}}
                if mo not in mo_param_dict[parent_mo]["children"]:
                    mo_param_dict[parent_mo]["children"].append(mo)

            param_entry = {
                "type": row["param_type"],
                "default": row["default_value"],
                "required": row["required"],
                "range": row["range"]
            }

            if parent:
                if parent not in mo_param_dict[mo]["params"] or \
                        mo_param_dict[mo]["params"][parent].get("type") != "list":
                    mo_param_dict[mo]["params"][parent] = {
                        "type": "list",
                        "isList": True,
                        "children": {}
                    }
                mo_param_dict[mo]["params"][parent]["children"][pname] = param_entry

            elif row["param_type"] != "structure" and int(row.get("multiplicity", "1")) >= 2:
                mo_param_dict[mo]["params"][pname] = {
                    "type": "list",
                    "isList": True,
                    "children": {
                        "val": param_entry
                    }
                }

            else:
                mo_param_dict[mo]["params"][pname] = param_entry

        version = "unknown"
        for part in excel_filename.split("_"):
            if part.startswith("24") or part.startswith("25"):
                version = part
                break

        output_filename = f"{version}_mo_param_dict.json"
        output_path = os.path.join(data_dir, output_filename)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(mo_param_dict, f, indent=2, ensure_ascii=False)

        self.poutput(f"MO 파라미터 dict가 저장되었습니다: {output_path}")

    def _parse_scf_to_dict(self, xml_path):
        tree = ET.parse(xml_path)
        root = tree.getroot()

        mo_dict = {}
        for mo in root.iter("managedObject"):
            dist = mo.attrib.get("distName")
            params = {p.attrib["name"]: p.text for p in mo.findall("p")}
            mo_dict[dist] = params
        return mo_dict

    def _compare_scf_dicts(self,dict_a, dict_b):
        """
        dict_a → dict_b 로의 변화 분석
        """
        only_in_b = sorted(set(dict_b) - set(dict_a))
        only_in_a = sorted(set(dict_a) - set(dict_b))
        changed = []

        for dist in set(dict_a) & set(dict_b):
            diff = {}
            for k, v in dict_a[dist].items():
                v2 = dict_b[dist].get(k)
                if v2 is not None and v2 != v:
                    diff[k] = (v, v2)
            if diff:
                changed.append((dist, diff))

        return only_in_b, only_in_a, changed

    def do_compare_scf(self, arg):
        """
        두 개의 SCF(XML) 파일을 비교하여 차이점을 출력합니다.
        사용법: compare-scf <파일 A>.xml <파일 B>.xml
        (파일은 data/scf/ 폴더 아래 있어야 하며 A → B 기준으로 비교됩니다)
        """
        tokens = arg.strip().split()
        if len(tokens) != 2:
            self.perror("사용법: compare-scf <파일 A>.xml <파일 B>.xml")
            return

        file_a, file_b = tokens
        base_dir = os.path.dirname(os.path.abspath(__file__))
        scf_dir = os.path.join(base_dir, "..", "data", "generated")

        file_a_path = os.path.join(scf_dir, file_a)
        file_b_path = os.path.join(scf_dir, file_b)

        if not os.path.exists(file_a_path):
            self.perror(f"[오류] 파일 없음: {file_a_path}")
            return
        if not os.path.exists(file_b_path):
            self.perror(f"[오류] 파일 없음: {file_b_path}")
            return

        try:
            dict_a = self._parse_scf_to_dict(file_a_path)
            dict_b = self._parse_scf_to_dict(file_b_path)
            only_in_b, only_in_a, changed = self._compare_scf_dicts(dict_a, dict_b)

            self.poutput(f"[비교 기준] {file_a} → {file_b}")
            self.poutput("")

            self.poutput(f"[➕ {file_b} 에만 있는 MO]")
            if only_in_b:
                for dist in only_in_b:
                    self.poutput(f"+ {dist}")
            else:
                self.poutput("(없음)")
            self.poutput("")

            self.poutput(f"[➖ {file_a} 에만 있던 MO]")
            if only_in_a:
                for dist in only_in_a:
                    self.poutput(f"- {dist}")
            else:
                self.poutput("(없음)")
            self.poutput("")

            self.poutput(f"[✏ 변경된 MO (파라미터 값 변경: {file_a} → {file_b})]")
            if changed:
                for dist, diff in changed:
                    self.poutput(f"* {dist}")
                    for k, (v1, v2) in diff.items():
                        self.poutput(f"    - {k}: {v1} → {v2}")
            else:
                self.poutput("(없음)")

        except Exception as e:
            self.perror(f"[오류] 비교 중 예외 발생: {e}")
