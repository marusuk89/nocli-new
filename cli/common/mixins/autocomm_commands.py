import base64
import os
import shlex
import json
import openpyxl
import re
import cli.core.config_x.constants as constans
from io import BytesIO
from collections import OrderedDict
from datetime import datetime
from cli.settings import is_debug
from collections import OrderedDict
from cli.common.util.tmpl_utils import load_cablink_blocks, load_prod_code_maps
from cli.common.util.server_utils import load_from_server, save_to_server, delete_from_server
from proto import message_pb2
from cli.settings import grpc_stub
from cli.common.util.path_utils import get_path

class AutocommCommandMixin:
    def do_show_script_header(self, arg):
        """
        스크립트 파일의 HEADER 영역 명령어들을 서버에서 받아와 출력합니다.
        사용법: show-script-header <파일명>
        """
        filename = arg.strip()
        if not filename:
            self.perror("사용법: show-script-header <파일명>")
            return

        try:
            # 서버로부터 스크립트 내용 로딩 (text 모드, autocomm 용도)
            script_text = load_from_server(filename, filetype="text", purpose="autocomm")

            if script_text is None:
                self.perror(f"[오류] 서버로부터 파일을 받아오지 못했습니다: {filename}")
                return

            in_header = False
            header_lines = []

            for line in script_text.splitlines():
                stripped = line.strip()

                if stripped == "### HEADER ###":
                    in_header = True
                    continue
                elif stripped == "### BODY ###":
                    break

                if in_header:
                    header_lines.append(line.rstrip())

            if header_lines:
                self.poutput("\n".join(header_lines))
            else:
                self.poutput("[안내] HEADER 구간이 비어있거나 존재하지 않습니다.")

        except Exception as e:
            self.perror(f"[오류] show-script-header 실패: {e}")


    def do_autocomm_run_script(self, arg):
        """
        서버에서 받은 스크립트 파일의 BODY 영역 명령어들을 실행합니다.
        사용법: autocomm-run-script <파일명>
        """
        self.xml_tree = self._create_empty_xml()
        filename = arg.strip()
        if not filename:
            self.perror("사용법: autocomm-run-script <파일명>")
            return

        try:
            # 서버로부터 스크립트 전체 텍스트 로드
            script_text = load_from_server(filename, filetype="text", purpose="autocomm")

            if script_text is None:
                self.perror(f"[오류] 서버로부터 파일을 받아오지 못했습니다: {filename}")
                return

            in_body = False  # BODY 시작 여부 플래그
            for i, raw_line in enumerate(script_text.splitlines(), 1):
                line = raw_line.strip()

                if line == "### BODY ###":
                    in_body = True
                    continue
                elif line == "### HEADER ###":
                    in_body = True
                    continue

                if not in_body or not line or line.startswith("#"):
                    continue

                if is_debug:
                    print(f"[디버그] {i}번째 줄 읽음: {repr(line)}")

                self.last_script_line = line
                tokens = shlex.split(line)
                if not tokens:
                    continue

                cmd = tokens[0]
                cmd_func = getattr(self, f"do_{cmd.replace('-', '_')}", None)

                if is_debug:
                    print("cmd       :", cmd)
                    print("cmd_func  :", cmd_func)

                if callable(cmd_func):
                    normalized_cmd = cmd.replace("-", "_")
                    new_line = " ".join([normalized_cmd] + tokens[1:])

                    if is_debug:
                        print(f"[실행] 공식 명령어 실행 → {normalized_cmd}")
                        print(f"new_line : {new_line}")

                    self._original_line = line
                    self.onecmd(new_line)
                else:
                    if is_debug:
                        print(f"[실행] default로 처리 → {line}")
                    self._original_line = line
                    self.default(line)
            if is_debug:
                self.poutput("[완료] autocomm-run-script 실행 완료.")

        except Exception as e:
            self.perror(f"[오류] autocomm-run-script 실패: {e}")

    def do_list_tmpl(self, arg):
        """
        날짜별 엑셀 템플릿 목록을 출력합니다.
        사용법: list-tmpl <YYYYMMDD>
        """
        date_str = arg.strip()
        if not date_str:
            self.perror("사용법: list-tmpl <YYYYMMDD>")
            return

        try:
            datetime.strptime(date_str, "%Y%m%d")
        except ValueError:
            self.perror("날짜 형식이 올바르지 않습니다. 예: 20250703")
            return

        payload = date_str
        request = message_pb2.Request(command="listTmpl", payload=payload)

        try:
            response = grpc_stub.SendCommand(request)
        except Exception as e:
            self.perror(f"[오류] 서버 통신 실패: {e}")
            return

        if not response.success:
            self.perror(f"[서버 오류] {response.result}")
            return

        xlsx_files = response.result.split("||") if response.result else []

        if not xlsx_files:
            self.poutput(f"[정보] 해당 날짜({date_str})의 엑셀 템플릿이 없습니다.")
            return

        for f in sorted(xlsx_files):
            self.poutput(f"> {f}")

    def do_list_script(self, arg):
        """
        날짜별 CLI 스크립트 목록을 서버에서 받아 출력합니다.
        사용법: list-script <YYYYMMDD>
        """
        date_str = arg.strip()
        if not date_str:
            self.perror("사용법: list-script <YYYYMMDD>")
            return

        try:
            datetime.strptime(date_str, "%Y%m%d")
        except ValueError:
            self.perror("날짜 형식이 잘못되었습니다. 예: 20250623")
            return
        
        payload = date_str
        request = message_pb2.Request(command="listScript", payload=payload)

        try:
            response = grpc_stub.SendCommand(request)
        except Exception as e:
            self.perror(f"[오류] 서버 통신 실패: {e}")
            return

        if not response.success:
            self.perror(f"[서버 오류] {response.result}")
            return

        file_list = response.result.split("||") if response.result else []
        if not file_list:
            self.poutput("[정보] 해당 날짜 디렉토리에 CLI 스크립트가 없습니다.")
            return

        for filename in sorted(file_list):
            self.poutput(f"> {filename}")

    def do_gen_script(self, arg):
        """
        INIT-BTS / INIT-CELL CLI 스크립트를 엑셀에서 생성합니다.
        사용법: gen-script <엑셀파일명>
        """
        excel_filename = arg.strip()
        if not excel_filename:
            self.perror("사용법: gen-script <엑셀파일명>")
            return

        excel_bytes = load_from_server(filename=excel_filename, filetype="binary", purpose="autocomm")
        if not excel_bytes:
            self.perror(f"서버에서 엑셀 파일을 불러오지 못했습니다: {excel_filename}")
            return

        try:
            wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)

            # "INITBTS" 포함 시트 찾기
            ws_bts = None
            for sheet_name in wb.sheetnames:
                if "INITBTS" in sheet_name.upper():
                    ws_bts = wb[sheet_name]
                    break
            if ws_bts is None:
                raise KeyError("INITBTS 시트를 찾을 수 없습니다.")

            # "INITCELL" 포함 시트 찾기
            ws_cell = None
            for sheet_name in wb.sheetnames:
                if "INITCELL" in sheet_name.upper():
                    ws_cell = wb[sheet_name]
                    break
            if ws_cell is None:
                raise KeyError("INITCELL 시트를 찾을 수 없습니다.")

            # MOD 시트도 부분 일치로 찾을 수 있게 변경
            ws_mod = None
            for sheet_name in wb.sheetnames:
                if "INITMOD" in sheet_name.upper():
                    ws_mod = wb[sheet_name]
                    break
            if ws_mod is None:
                raise KeyError("INITMOD 시트를 찾을 수 없습니다.")

        except KeyError as e:
            self.perror(f"엑셀 시트 오류: {e}")
            return
        except Exception as e:
            self.perror(f"엑셀 파일을 열 수 없습니다: {e}")
            return


        # 실제 파싱은 내부 함수에서 처리하므로 여기서는 단순 위임
        script_entries_bts = self._parse_excel_rows_init_bts(ws_bts)
        script_entries_cell = self._parse_excel_rows_init_cell(ws_cell)
        script_entries_mod = self._parse_excel_rows_init_mod(ws_mod)

        self._write_script_files(script_entries_bts, script_entries_cell, script_entries_mod)

    def _parse_excel_rows_init_bts(self, ws):
        script_entries = []
        prod_map = load_prod_code_maps()
        today_str = datetime.now().strftime("%Y%m%d")

        scenario_row_idx = self._find_scenario_row_bts(ws)
        if scenario_row_idx is None:
            self.perror("엑셀 시트에서 'CLI SCENARIO' 셀을 찾을 수 없습니다.")
            return []

        mo_headers, param_keys, data_start_row_idx, max_col = self._parse_header_and_key_rows_bts(ws, scenario_row_idx)
        mo_columns = self._build_mo_columns_bts(mo_headers, param_keys)
        data_end_row_idx = self._find_data_end_row_bts(ws, data_start_row_idx)

        try:
            action_col = mo_headers.index("CLI Scenario")
        except ValueError:
            self.perror("'CLI Scenario' 열이 존재하지 않습니다.")
            return []

        for row_idx in range(data_start_row_idx, data_end_row_idx):
            fail = False
            error_lines = []
            row = [ws.cell(row=row_idx, column=c + 1).value for c in range(max_col)]

            action_val = row[action_col]
            if str(action_val).strip().upper() != "INIT-BTS":
                fail = True
                error_lines.append(f"action_val 값이 INIT-BTS가 아님 (row {row_idx + 1})")
                continue

            parsed = self._parse_row_values_bts(row, mo_columns)

            # 추가 검증: radio_type / radio_ver 유효성
            radio_type = parsed.get("radio_type")
            if radio_type and radio_type not in constans.VALID_RADIO_TYPES:
                fail = True
                error_lines.append(f"radio_type 값이 유효하지 않음: '{radio_type}' (row {row_idx + 1})")

            radio_ver = parsed.get("radio_ver")
            if radio_ver and radio_ver not in constans.VALID_RADIO_VERSIONS:
                fail = True
                error_lines.append(f"radio_ver 값이 유효하지 않음: '{radio_ver}' (row {row_idx + 1})")
            
            # 필수 필드 존재 여부 검증
            if radio_type.upper() == "4G":
                required_keys = [
                    ("mrbts_id", "MRBTS:id"),
                    ("enb_name", "LNBTS:enbName"),
                    ("prod_name", "SMOD:prod-name"),
                    ("ip_id", "IPIF:id"),
                    ("ip_addr", "IPIF:localIpAddr"),
                    ("prefix_len", "IPIF:localIpPrefixLength"),
                    ("iprt_gateway", "IPRT:gateway"),
                    ("vlanif_id", "VLANIF:id"),
                    ("vlan_id", "VLANIF:vlanId"),
                    ("sync_id_col", "SYNC-1:id"),
                    ("sync_prio_col", "SYNC-1:syncInputPrio"),
                    ("sync_type_col", "SYNC-1:syncInputType"),
                    ("ntp_id", "NTP:id"),
                    ("ntp_server_ip_addr", "NTP:ntpServerIpAddrList[]"),
                    ("netact_id", "CTRLTS-1:id"),
                    ("netact_ip_addr", "CTRLTS-1:netActIpAddr"),
                    ("radio_type", "Radio:Type"),
                    ("radio_ver", "Radio:Version"),
                ]
            elif radio_type.upper() == "5G":
                required_keys = [
                    ("mrbts_id", "MRBTS:id"),
                    ("enb_name", "LNBTS:enbName"),
                    ("prod_name", "SMOD:prod-name"),
                    ("ip_id", "IPIF:id"),
                    ("ip_addr", "IPIF:localIpAddr"),
                    ("prefix_len", "IPIF:localIpPrefixLength"),
                    ("iprt_gateway", "IPRT:gateway"),
                    ("vlanif_id", "VLANIF:id"),
                    ("vlan_id", "VLANIF:vlanId"),
                    ("sync_id_col", "SYNC-1:id"),
                    ("sync_prio_col", "SYNC-1:syncInputPrio"),
                    ("sync_type_col", "SYNC-1:syncInputType"),
                    ("netact_id", "CTRLTS-1:id"),
                    ("netact_ip_addr", "CTRLTS-1:netActIpAddr"),
                    ("radio_type", "Radio:Type"),
                    ("radio_ver", "Radio:Version"),
                ]
            for key, label in required_keys:
                if not parsed.get(key):
                    fail = True
                    error_lines.append(f"{label} 값이 누락됨 (row {row_idx + 1})")

            vlan_id = parsed.get("vlan_id")
            if vlan_id:
                try:
                    vlan_num = int(vlan_id)
                    if not (0 <= vlan_num <= 4094):
                        fail = True
                        error_lines.append(f"vlan_id 값이 유효한 범위(0~4094)가 아님: '{vlan_id}' (row {row_idx + 1})")
                except ValueError:
                    fail = True
                    error_lines.append(f"vlan_id 값이 정수가 아님: '{vlan_id}' (row {row_idx + 1})")

            # MTRACE ID/ADDR 개수 일치 여부
            id_count = len(parsed.get("ctrlts1_mtrace_ids", []))
            addr_count = len(parsed.get("ctrlts1_mtrace_addrs", []))
            if id_count != addr_count:
                fail = True
                error_lines.append(f"CTRLTS-1/MTRACE id({id_count})와 address({addr_count}) 개수 불일치 (row {row_idx + 1})")

            mrbts_id = parsed.get("mrbts_id") or f"row{row_idx+1}"
            du_type = self._determine_du_type_bts(parsed.get("prod_name"), radio_type)

            if fail:
                script = [
                    "# [ERROR] 이 CLI 스크립트는 실행되지 않습니다.",
                    "# 로그 파일을 확인하세요: parse_errors.log"
                ]
            else:
                script = self._build_script_bts(parsed, du_type, prod_map)

            script_entries.append((str(mrbts_id), script, "bts", fail, error_lines))

        return script_entries

    def _find_scenario_row_bts(self, ws):
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().upper() == "CLI SCENARIO":
                return row_idx
        return None


    def _parse_header_and_key_rows_bts(self, ws, scenario_row_idx):
        mo_row_idx = scenario_row_idx
        key_row_idx = mo_row_idx + 1
        data_start_row_idx = mo_row_idx + 2
        max_col = ws.max_column

        mo_headers = [ws.cell(row=mo_row_idx, column=c).value for c in range(1, max_col + 1)]
        param_keys = [ws.cell(row=key_row_idx, column=c).value for c in range(1, max_col + 1)]

        for i in range(len(mo_headers)):
            if mo_headers[i] is None and i > 0:
                mo_headers[i] = mo_headers[i - 1]

        return mo_headers, param_keys, data_start_row_idx, max_col


    def _build_mo_columns_bts(self, mo_headers, param_keys):
        mo_columns = {}
        current_mo = None
        for col in range(len(mo_headers)):
            mo = mo_headers[col]
            param = param_keys[col]
            if mo:
                current_mo = mo
                if current_mo not in mo_columns:
                    mo_columns[current_mo] = []
            if current_mo and param:
                mo_columns[current_mo].append((param, col))
        return mo_columns


    def _find_data_end_row_bts(self, ws, data_start_row_idx):
        for row_idx in range(data_start_row_idx, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().startswith("#####"):
                return row_idx
        return ws.max_row + 1


    def _determine_du_type_bts(self, prod_name, radio_type):
        if not isinstance(prod_name, str):
            return None
        prod_upper = prod_name.strip().upper()
        if prod_upper.startswith("AS") and radio_type.upper() == "4G":
            return "du20"
        elif prod_upper.startswith("FS") and radio_type.upper() == "4G":
            return "FSMF"
        elif prod_upper == "ASIK" and radio_type.upper() == "5G":
            return "du10"
        elif prod_upper == "ASIL" and radio_type.upper() == "5G":
            return "du20"
        return None

    def _parse_row_values_bts(self, row, mo_columns):
        parsed = {
            "ctrlts1_mtrace_ids": [],
            "ctrlts1_mtrace_addrs": [],
            "bbmod_list": [],
        }

        # (A) radio_type 선행 추출
        radio_type = None
        for mo, param_list in mo_columns.items():
            for param, col in param_list:
                if mo == "Radio" and param == "Type":
                    radio_type = row[col] if col < len(row) else None
                    break

        # (B) RAT별 MO 별칭 매핑 로드 (5G면 LNBTS→NRBTS)
        mo_alias = self._mo_alias_map_for_bts(radio_type)

        # (C) 본 파싱
        for mo, param_list in mo_columns.items():
            mo_canon = mo_alias.get(mo, mo)  # 논리적 MO명

            bbmod_id = bbmod_prod = None
            for param, col in param_list:
                val = row[col] if col < len(row) else ""
                key = f"{mo}:{param}"    # 원본 키도 그대로 보존(디버깅/검증용)
                parsed[key] = val

                if mo == "MRBTS" and param == "id":
                    parsed["mrbts_id"] = val

                # LNBTS/NRBTS 공통 이름 키로 enb_name만 받음 (빌더에서 실제 명령 변환)
                if mo in ("LNBTS",) or mo_canon in ("NRBTS",):
                    if param == "enbName":
                        parsed["enb_name"] = val.replace("_", "") if val else val

                elif mo == "SMOD" and param == "prod-name":
                    parsed["prod_name"] = val

                elif "IPIF" in mo:
                    if param == "id":
                        parsed["ip_id"] = val
                    elif param == "localIpAddr":
                        parsed["ip_addr"] = val
                    elif param == "localIpPrefixLength":
                        parsed["prefix_len"] = val

                elif "IPRT" in mo and param.endswith("gateway"):
                    parsed["iprt_gateway"] = val

                elif "VLANIF" in mo:
                    if param == "id":
                        parsed["vlanif_id"] = val
                    elif param == "vlanId":
                        parsed["vlan_id"] = val

                elif "SYNC-1" in mo:
                    if param == "id":
                        parsed["sync_id_col"] = val
                    elif param.endswith("syncInputPrio"):
                        parsed["sync_prio_col"] = val
                    elif param.endswith("syncInputType"):
                        parsed["sync_type_col"] = val

                elif "NTP" in mo:
                    if param == "id":
                        parsed["ntp_id"] = val
                    elif param == "ntpServerIpAddrList[]":
                        parsed["ntp_server_ip_addr"] = val

                elif mo == "BBMOD":
                    if param == "id":
                        bbmod_id = val
                    elif param == "prod-name":
                        bbmod_prod = val
                    if bbmod_id and bbmod_prod:
                        parsed["bbmod_list"].append((bbmod_id, bbmod_prod))

                elif mo.startswith("FHS_") and param == "id":
                    parsed[f"{mo.lower()}_id"] = val

                elif mo == "CTRLTS-1" and param == "id":
                    parsed["netact_id"] = val
                elif mo == "CTRLTS-1" and param == "netActIpAddr":
                    parsed["netact_ip_addr"] = val
                elif mo == "CTRLTS-1/MTRACE":
                    if param == "id":
                        parsed["ctrlts1_mtrace_ids"].append(val)
                    elif param == "tceIpAddress":
                        parsed["ctrlts1_mtrace_addrs"].append(val)

                elif mo == "Radio" and param == "Type":
                    parsed["radio_type"] = val
                elif mo == "Radio" and param == "Version":
                    parsed["radio_ver"] = val

        return parsed
    
    def _mo_alias_map_for_bts(self, radio_type: str) -> dict:
        """
        파싱 단계에서만 쓰는 얕은 매핑.
        - 5G일 때 엑셀의 LNBTS 값을 논리적으로 NRBTS로 본다.
        - 엑셀 시트는 공통이므로 실제 셀 라벨은 그대로 둔다.
        """
        rt = (radio_type or "").strip().upper()
        if rt == "5G":
            return {"LNBTS": "NRBTS"}
        return {}

    def _build_script_bts(self, parsed, du_type, prod_map):
        rt = (parsed.get("radio_type") or "").strip().upper()
        if rt == "5G":
            return self._build_script_bts_5g(parsed, du_type, prod_map)
        return self._build_script_bts_lte(parsed, du_type, prod_map)
    
    def _build_script_bts_lte(self, parsed, du_type, prod_map):
        mrbts_id = parsed.get("mrbts_id")
        radio_ver = parsed.get("radio_ver")
        radio_type = parsed.get("radio_type")
        enb_name = parsed.get("enb_name")
        prod_name = parsed.get("prod_name")

        ip_id = parsed.get("ip_id")
        ip_addr = parsed.get("ip_addr")
        prefix_len = parsed.get("prefix_len")
        iprt_gateway = parsed.get("iprt_gateway")

        vlanif_id = parsed.get("vlanif_id")
        vlan_id = parsed.get("vlan_id")

        sync_id_col = parsed.get("sync_id_col")
        sync_prio_col = parsed.get("sync_prio_col")
        sync_type_col = parsed.get("sync_type_col")
        ntp_id = parsed.get("ntp_id")
        ntp_server_ip_addr = parsed.get("ntp_server_ip_addr")

        bbmod_list = parsed.get("bbmod_list", [])
        netact_id = parsed.get("netact_id")
        netact_ip_addr = parsed.get("netact_ip_addr")
        ctrlts1_mtrace_ids = parsed.get("ctrlts1_mtrace_ids", [])
        ctrlts1_mtrace_addrs = parsed.get("ctrlts1_mtrace_addrs", [])

        script = ["### HEADER ###"]
        script += [
            f"set-bts {mrbts_id} {ip_id}",
            f"dest-bts {mrbts_id}",
            f"check-ping {mrbts_id}",
            f"set-mode bts",
            f"set-mo-version {radio_ver}",
            f"set-rat-type {radio_type}",
            f"set-iot-lncel-id 65534",
            "set-allow-commit-diff true",
            ""
        ]

        script += [
            "### BODY ###",
            f"set-du-type {du_type}",
            f"tgt-bts {mrbts_id}",
            f"exec-script {radio_ver}_{du_type.upper()}_INITBTS_FILTER_SCRIPT_V0.7.5.cli",
            "# 진입 단계 종료", "",
            f"btsName {enb_name}", "# MRBTS 단계 종료", "",
            f"LNBTS {mrbts_id}", f"enbName {enb_name}",
            "exit-all", "# LNBTS 단계 종료", ""
        ]

        if du_type.upper() == "FSMF":
            script += ["# 더미 설정 시작", "EQM 1", "HWTOP 1", "CABLINK 65535",
                    f"firstEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/CABINET-1/SMOD-1",
                    "firstEndpointLabel  OPT", "firstEndpointPortId  6",
                    f"secondEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/RMOD-32767",
                    "secondEndpointLabel  OPT", "secondEndpointPortId  1",
                    "exit-all",
                    f"LNBTS {mrbts_id}", "LNCEL 65535", "lcrId 255", "nbIotLinkedCellId 254", "exit",
                    "LNCEL 65534", "lcrId 254", "nbIotLinkedCellId 255", "exit-all",
                    "#더미 설정 끝", ""]
        elif du_type.upper == "DU20":
            script += ["# 더미 설정 시작", "EQM 1", "HWTOP 1", "CABLINK 65535",
                    f"firstEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/FHS-1",
                    "firstEndpointLabel  OPT", "firstEndpointPortId  12",
                    f"secondEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/RMOD-32767",
                    "secondEndpointLabel  OPT", "secondEndpointPortId  1",
                    "exit-all",
                    f"LNBTS {mrbts_id}", "LNCEL 65535", "lcrId 255", "nbIotLinkedCellId 254", "exit",
                    "LNCEL 65534", "lcrId 254", "nbIotLinkedCellId 255", "exit-all",
                    "#더미 설정 끝", ""]

        if prod_name:
            mapped = prod_map.get(prod_name.upper(), prod_name)
            script += ["EQM 1", "APEQM 1", "CABINET 1", "SMOD 1", f"prodCodePlanned {mapped}", "exit-all", "# SMOD 단계 종료", ""]

        if ip_id and ip_addr and prefix_len and iprt_gateway:
            script += ["# TNLSVC → IPIF 설정", "TNLSVC 1", "TNL 1", "IPNO 1", "IPIF 1",
                    f"IPADDRESSV4 {ip_id}", f"localIpAddr {ip_addr}",
                    f"localIpPrefixLength {prefix_len}", "exit", "exit",
                    f"IPRT {ip_id}", f"list staticRoutes {ip_id} gateway {iprt_gateway}",
                    f"list staticRoutes {ip_id} routeIpMtu 1510",
                    f"list staticRoutes {ip_id} destIpAddr 0.0.0.0",
                    "exit-all", "# TNLSVC 설정 종료", ""]

        if vlanif_id and vlan_id:
            script += ["# TNLSVC → ETHSVC 설정", "TNLSVC 1", "TNL 1", "ETHSVC 1", "ETHIF 1",
                    f"VLANIF {vlanif_id}", f"vlanId {vlan_id}",
                    "exit-all", "# ETHSVC 설정 종료", ""]

        if sync_id_col and sync_prio_col and sync_type_col:
            SYNC_TYPE_MAP = {
                "Master": "1pps/ToD from external GNSS receiver",
                "Slave": "1pps/ToD from Sync Hub Master",
                "Backplane": "1pps/ToD from backplane",
                "TOPP": "TOPP"
            }
            sync_type_value = SYNC_TYPE_MAP.get(str(sync_type_col).strip(), sync_type_col)
            script += ["# SYNC-1/CLOCK 구성", "MNL 1", "MNLENT 1", "SYNC 1",
                    f"CLOCK {sync_id_col}",
                    f"list syncInputList 1 syncInputPrio {sync_prio_col}",
                    f'list syncInputList 1 syncInputType "{sync_type_value}"']
            if sync_type_value == "1pps/ToD from external GNSS receiver":
                script += ["GNSSE 1", "exit"]
            script += [f"NTP {ntp_id}"]
            ntp_ips = [ip.strip() for ip in str(ntp_server_ip_addr).split(";") if ip.strip()]
            for i, ip in enumerate(ntp_ips, 1):
                key = "ntpServerIpAddrList" if du_type == "FSMF" else "ntpServerIpAddrOrFqdnList"
                script += [f"list {key} {i} val {ip}"]
            script += ["maxNtpTimeError 200", "ntpAlarmingThreshold 300", "ntpDscp 46", "exit-all", ""]

        if bbmod_list:
            script += ["# BBMOD 설정", "EQM 1", "APEQM 1", "CABINET 1"]
            for bb_id, bb_prod in bbmod_list:
                mapped = prod_map.get(bb_prod.upper(), bb_prod)
                script += [f"BBMOD {bb_id}", f"prodCodePlanned {mapped}"]
                if bb_prod == "FBBA":
                    script += ["no-list srioConnectionList"]
                script += ["exit"]
            script += ["exit-all"]

        if du_type == "du20" and bbmod_list:
            script += ["# FHS 설정", "EQM 1", "APEQM 1"]
            for i in range(1, 7):
                if parsed.get(f"fhs_{i}_id"):
                    script += [f"FHS {i}", "exit"]
            script += ["exit-all", "# FHS 설정 끝", ""]

        if netact_id and netact_ip_addr:
            script += [f"LNBTS {mrbts_id}", f"CTRLTS {netact_id}", f"netActIpAddr {netact_ip_addr}"]

        for m_id, m_addr in zip(ctrlts1_mtrace_ids, ctrlts1_mtrace_addrs):
            if m_id is not None and m_addr:
                script += [f"MTRACE {m_id}", f"tceIpAddress {m_addr}", "exit"]
        script += ["exit-all"]

        script += ["commit-all -t 1", "apply-bts-cfg-commission", "#act-bts-cfg-commission"]

        return script
    
    def _build_script_bts_5g(self, parsed, du_type, prod_map):
        """
        5G 전용 init-bts 스크립트 빌더 (초기 골격).
        - HEADER/BODY 프레임과 공통 블록 호출 중심.
        - NRBTS/NRCELL 등 NR 고유 블록은 다음 단계에서 정확한 명령으로 채운다.
        """
        mrbts_id = parsed.get("mrbts_id")
        radio_ver = parsed.get("radio_ver")
        radio_type = parsed.get("radio_type")
        enb_name = parsed.get("enb_name")
        prod_name = parsed.get("prod_name")

        ip_id = parsed.get("ip_id")
        ip_addr = parsed.get("ip_addr")
        prefix_len = parsed.get("prefix_len")
        iprt_gateway = parsed.get("iprt_gateway")

        vlanif_id = parsed.get("vlanif_id")
        vlan_id = parsed.get("vlan_id")

        sync_id_col = parsed.get("sync_id_col")
        sync_prio_col = parsed.get("sync_prio_col")
        sync_type_col = parsed.get("sync_type_col")

        bbmod_list = parsed.get("bbmod_list", [])
        netact_id = parsed.get("netact_id")
        netact_ip_addr = parsed.get("netact_ip_addr")
        ctrlts1_mtrace_ids = parsed.get("ctrlts1_mtrace_ids", [])
        ctrlts1_mtrace_addrs = parsed.get("ctrlts1_mtrace_addrs", [])

        script = ["### HEADER ###"]
        script += [
            f"set-bts {mrbts_id} {ip_id}",
            f"dest-bts {mrbts_id}",
            f"check-ping {mrbts_id}",
            f"set-mode bts",
            f"set-mo-version {radio_ver}",
            f"set-rat-type {radio_type}",
            "set-allow-commit-diff true",
            ""
        ]

        script += ["### BODY ###"]
        script += [
            f"set-du-type {du_type}",
            f"tgt-bts {mrbts_id}",
            f"exec-script DU00_5G_INITBTS_FILTER_SCRIPT_V1.0.cli",
            f"exec-script DU10_5G_INITBTS_FILTER_SCRIPT_V1.0.cli",
        ]

        # enb_name이 NFO_로 시작하면 추가 실행
        if enb_name and enb_name.startswith("NFO_"):
            script.append(f"exec-script DU00_5G_INITBTS_FILTER_SCRIPT_NFO_V1.0.cli")

        script += [
            "# 진입 단계 종료", "",
            f"btsName {enb_name}", "# MRBTS 단계 종료", "",
            f"NRBTS {mrbts_id}",
            "exit-all", "# LNBTS 단계 종료", ""
            ""
        ]

        if du_type.upper() == "DU10":
            script += ["# 더미 설정 시작", "EQM 1", "HWTOP 1", "CABLINK 65535",
                    f"firstEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/CABINET-1/BBMOD-1",
                    "firstEndpointLabel  SFP", "firstEndpointPortId  1",
                    'iqCompression "eCPRI IQ9_E7"',
                    "linkSpeed eCpri10",
                    "radioProtocolType 10eCPRI",
                    f"secondEndpointDN  MRBTS-{mrbts_id}/EQM-1/APEQM-1/RMOD-31",
                    "secondEndpointLabel  SFP", "secondEndpointPortId  1",
                    "exit-all",
                    f"NRBTS {mrbts_id}", "NRCELL 6144", "lcrId 31", "exit-all",
                    "#더미 설정 끝", ""]
        elif du_type.upper() == "DU20":
            script += ["# 더미 설정 시작",
                    "#더미 설정 끝", ""]
            
        if prod_name:
            mapped = prod_map.get(prod_name.upper(), prod_name)
            script += ["EQM 1", "APEQM 1", "CABINET 1", "SMOD 1", f"prodCodePlanned {mapped}", "exit-all", "# SMOD 단계 종료", ""]

        if ip_id and ip_addr and prefix_len and iprt_gateway:
            script += ["# TNLSVC → IPIF 설정", "TNLSVC 1", "TNL 1", "IPNO 1", "IPIF 1",
                    f"IPADDRESSV4 {ip_id}", f"localIpAddr {ip_addr}",
                    f"localIpPrefixLength {prefix_len}", "exit", "exit",
                    f"IPRT {ip_id}", f"list staticRoutes {ip_id} gateway {iprt_gateway}",
                    "exit-all", "# TNLSVC 설정 종료", ""]
            
        if vlanif_id and vlan_id:
            script += ["# TNLSVC → ETHSVC 설정", "TNLSVC 1", "TNL 1", "ETHSVC 1", "ETHIF 1",
                    f"VLANIF {vlanif_id}", f"vlanId {vlan_id}",
                    "exit-all", "# ETHSVC 설정 종료", ""]
            
        if sync_id_col and sync_prio_col and sync_type_col:
            SYNC_TYPE_MAP = {
                "Master": "1pps/ToD from external GNSS receiver",
                "Slave": "1pps/ToD from Sync Hub Master",
                "Backplane": "1pps/ToD from backplane",
                "TOPP": "TOPP"
            }
            sync_type_value = SYNC_TYPE_MAP.get(str(sync_type_col).strip(), sync_type_col)
            script += ["# SYNC-1/CLOCK 구성", "MNL 1", "MNLENT 1", "SYNC 1",
                    f"CLOCK {sync_id_col}",
                    f"list syncInputList 1 syncInputPrio {sync_prio_col}",
                    f'list syncInputList 1 syncInputType "{sync_type_value}"']
            if sync_type_value == "1pps/ToD from external GNSS receiver":
                script += ["GNSSE 1",]
            script += ["exit-all", ""]
            
        if bbmod_list:
            script += ["# BBMOD 설정", "EQM 1", "APEQM 1", "CABINET 1"]
            for bb_id, bb_prod in bbmod_list:
                mapped = prod_map.get(bb_prod.upper(), bb_prod)
                script += [f"BBMOD {bb_id}", f"prodCodePlanned {mapped}"]
                script += ["exit"]
            script += ["exit-all"]

        # if netact_id and netact_ip_addr:
        #     script += [f"NRBTS {mrbts_id}", f"CTRLTS {netact_id}", f"netActIpAddr {netact_ip_addr}"]

        for m_id, m_addr in zip(ctrlts1_mtrace_ids, ctrlts1_mtrace_addrs):
            if m_id is not None and m_addr:
                script += ["MNL 1", "TCEADM 1" , "TCE 1" ,f"tceIpAddress {m_addr}", "exit-all"]
        script += ["exit-all"]

        script += ["commit", "apply-bts-cfg-commission", "#act-bts-cfg-commission"]

        return script
    
    def _parse_excel_rows_init_cell(self, ws):
        script_entries = []
        scenario_row_idx = self._find_scenario_row_cell(ws)
        if scenario_row_idx is None:
            self.perror("엑셀 시트에서 'CLI SCENARIO' 셀을 찾을 수 없습니다.")
            return []

        mo_headers, param_keys, data_start_row_idx, max_col = self._parse_header_and_key_rows_cell(ws, scenario_row_idx)
        mo_columns = self._build_mo_columns_cell(mo_headers, param_keys)
        data_end_row_idx = self._find_data_end_row_cell(ws, data_start_row_idx)

        try:
            action_col = mo_headers.index("CLI Scenario")
        except ValueError:
            self.perror("'CLI Scenario' 열이 존재하지 않습니다.")
            return []

        du_dummy_killed = set()

        for row_idx in range(data_start_row_idx, data_end_row_idx):
            fail = False
            error_lines = []
            row = [ws.cell(row=row_idx, column=c + 1).value for c in range(max_col)]

            action_val = row[action_col]
            if not action_val or str(action_val).strip().upper() != "INIT-CELL":
                continue

            parsed = self._parse_row_values_cell(row, mo_columns)

            # radio_type / radio_ver 유효성
            radio_type = parsed.get("radio_type")
            if radio_type and radio_type.upper() not in constans.VALID_RADIO_TYPES:
                fail = True
                error_lines.append(f"radio_type 값이 유효하지 않음: '{radio_type}' (row {row_idx + 1})")

            radio_ver = parsed.get("radio_ver")
            if radio_ver and radio_ver.upper() not in constans.VALID_RADIO_VERSIONS:
                fail = True
                error_lines.append(f"radio_ver 값이 유효하지 않음: '{radio_ver}' (row {row_idx + 1})")

            if radio_type.upper() == "4G" :
                required_keys = [
                    ("mrbts_id", "MRBTS:id"),
                    ("mrbts_ip", "MRBTS:ip-addr"),
                    ("rmod_id", "RMOD:id"),
                    ("prod_name", "RMOD:prod-name"),
                    ("radio_type", "Radio:Type"),
                    ("radio_ver", "Radio:Version"),

                    ("lncel_id", "LNCEL/NRCELL:id"),
                    ("cell_name", "LNCEL/NRCELL:cellName"),
                    ("lcr_id", "LNCEL/NRCELL:lcrId"),
                    ("phy_cell_id", "LNCEL/NRCELL:phyCellId"),
                    ("tac", "LNCEL/NRCELL:tac"),
                    ("nbiot_linked_cell_id", "LNCEL/NRCELL:nbIotLinkedCellId"),

                    ("lncel_fdd_id", "LNCEL_FDD/NRCELL:id"),
                    ("root_seq_index", "LNCEL_FDD/NRCELL:rootSeqIndex"),

                    ("cablink_id", "CABLINK:id"),
                    ("dest_mod", "CABLINK:dest-mod"),
                    ("first_endpoint_label", "CABLINK:firstEndpointLabel"),
                    ("first_endpoint_port_id", "CABLINK:firstEndpointPortId"),
                    ("second_endpoint_label", "CABLINK:secondEndpointLabel"),
                    ("second_endpoint_port_id", "CABLINK:secondEndpointPortId"),
                ]
            elif radio_type.upper() == "5G" :
                required_keys = [
                    ("mrbts_id", "MRBTS:id"),
                    ("mrbts_ip", "MRBTS:ip-addr"),
                    ("rmod_id", "RMOD:id"),
                    ("prod_name", "RMOD:prod-name"),
                    ("radio_type", "Radio:Type"),
                    ("radio_ver", "Radio:Version"),

                    ("lncel_id", "LNCEL/NRCELL:id"),
                    ("cell_name", "LNCEL/NRCELL:cellName"),
                    ("lcr_id", "LNCEL/NRCELL:lcrId"),
                    ("phy_cell_id", "LNCEL/NRCELL:phyCellId"),
                    ("tac", "LNCEL/NRCELL:tac"),

                    ("root_seq_index", "LNCEL_FDD/NRCELL:rootSeqIndex"),

                    ("cablink_id", "CABLINK:id"),
                    ("dest_mod", "CABLINK:dest-mod"),
                    ("first_endpoint_label", "CABLINK:firstEndpointLabel"),
                    ("first_endpoint_port_id", "CABLINK:firstEndpointPortId"),
                    ("second_endpoint_label", "CABLINK:secondEndpointLabel"),
                    ("second_endpoint_port_id", "CABLINK:secondEndpointPortId"),
                ]

            for key, label in required_keys:
                val = parsed.get(key)

                # 특정 필드들은 0도 유효값으로 인정
                if key in ("lncel_fdd_id", "nbiot_linked_cell_id"):
                    if val is None or val == "":
                        fail = True
                        error_lines.append(f"{label} 값이 누락됨 (row {row_idx + 1})")
                else:
                    # 기본 로직: 0, None, "" 모두 누락 처리
                    if not val:
                        fail = True
                        error_lines.append(f"{label} 값이 누락됨 (row {row_idx + 1})")

            
            # 조건부 필수값 처리 (IOT 관련 필드)
            iot_required_prodnames = ["AHCA", "FXCA", "FRCG", "FHCA", "FHCB", "FHCG"]
            prod_name = str(parsed.get("prod_name") or "").strip().upper()

            if radio_type.upper() == "4G":
                iot_keys = [
                    ("iot_lncel_id", "LNCEL_IOT:id"),
                    ("iot_lcr_id", "LNCEL_IOT:lcrId"),
                    ("iot_nbiot_linked_cell_id", "LNCEL_IOT:nbIotLinkedCellId"),
                ]
                if prod_name in iot_required_prodnames:
                    for key, label in iot_keys:
                        if not parsed.get(key):
                            fail = True
                            error_lines.append(f"{label} 값이 누락됨 (prod_name={prod_name}, row {row_idx + 1})")
                else:
                    for key, label in iot_keys:
                        if parsed.get(key):  # 없어야 정상
                            fail = True
                            error_lines.append(f"{label} 값이 없어야 함 (prod_name={prod_name}, row {row_idx + 1})")

            # 결과 구성
            mrbts_id_str = str(parsed.get("mrbts_id") or f"row{row_idx+1}")
            mrbts_ip_str = str(parsed.get("mrbts_ip") or "UNKNOWN")
            lncel_id_str = str(parsed.get("lncel_id") or "UNKNOWN")

            prod_map = load_prod_code_maps()
            du_type = self._determine_du_type_bts(parsed.get("prod_name"), radio_type)

            if fail:
                script = [
                    "# [ERROR] 이 CLI 스크립트는 실행되지 않습니다.",
                    "# 로그 파일을 확인하세요: parse_errors.log"
                ]
            else:
                self.do_set_bts(f"{mrbts_id_str} {mrbts_ip_str}")
                self.do_dest_bts(mrbts_id_str)
                self.prepare_dummy_flag(mrbts_id_str, radio_type)
                script = self._build_script_cell(parsed, du_type, prod_map, du_dummy_killed)

            script_entries.append((mrbts_id_str, script, f"cell-{lncel_id_str}", fail, error_lines))

        return script_entries
    
    
    def _find_scenario_row_cell(self, ws):
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().upper() == "CLI SCENARIO":
                return row_idx
        return None

    def _parse_header_and_key_rows_cell(self, ws, scenario_row_idx):
        mo_row_idx = scenario_row_idx
        key_row_idx = mo_row_idx + 1
        data_start_row_idx = mo_row_idx + 2
        max_col = ws.max_column

        mo_headers = [ws.cell(row=mo_row_idx, column=c).value for c in range(1, max_col + 1)]
        param_keys = [ws.cell(row=key_row_idx, column=c).value for c in range(1, max_col + 1)]

        for i in range(len(mo_headers)):
            if mo_headers[i] is None and i > 0:
                mo_headers[i] = mo_headers[i - 1]

        return mo_headers, param_keys, data_start_row_idx, max_col


    def _build_mo_columns_cell(self, mo_headers, param_keys):
        mo_columns = {}
        current_mo = None
        for col in range(len(mo_headers)):
            mo = mo_headers[col]
            param = param_keys[col]
            if mo:
                current_mo = mo
                if current_mo not in mo_columns:
                    mo_columns[current_mo] = []
            if current_mo and param:
                mo_columns[current_mo].append((param, col))
        return mo_columns


    def _find_data_end_row_cell(self, ws, data_start_row_idx):
        for row_idx in range(data_start_row_idx, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().startswith("#####"):
                return row_idx
        return ws.max_row + 1
    
    def _parse_row_values_cell(self, row, mo_columns):
        parsed = {}
        for mo, key_list in mo_columns.items():
            for param, col in key_list:
                val = row[col] if col < len(row) else None
                key = f"{mo}:{param}"
                parsed[key] = val

                if mo == "MRBTS" and param == "id":
                    parsed["mrbts_id"] = val
                elif mo == "MRBTS" and param == "ip-addr":
                    parsed["mrbts_ip"] = val
                elif mo == "RMOD" and param == "id":
                    parsed["rmod_id"] = val
                elif mo == "RMOD" and param == "prod-name":
                    parsed["prod_name"] = val
                elif mo == "RMOD" and param == "cell-type":
                    parsed["cell_type"] = val
                elif mo == "RMOD" and param == "3-sector":
                    parsed["sector_3"] = val
                elif mo == "CABLINK":
                    if param == "id":
                        parsed["cablink_id"] = val
                    elif param == "dest-mod":
                        parsed["dest_mod"] = str(val).upper() if val else None
                    elif param == "firstEndpointLabel":
                        parsed["first_endpoint_label"] = val
                    elif param == "firstEndpointPortId":
                        parsed["first_endpoint_port_id"] = val
                    elif param == "secondEndpointLabel":
                        parsed["second_endpoint_label"] = val
                    elif param == "secondEndpointPortId":
                        parsed["second_endpoint_port_id"] = val
                elif mo == "LNCEL/NRCELL":
                    if param == "id":
                        parsed["lncel_id"] = val
                    elif param == "cellName":
                        parsed["cell_name"] = val
                    elif param == "lcrId":
                        parsed["lcr_id"] = val
                    elif param == "phyCellId":
                        parsed["phy_cell_id"] = val
                    elif param == "tac":
                        parsed["tac"] = val
                    elif param == "nbIotLinkedCellId":
                        parsed["nbiot_linked_cell_id"] = val
                elif mo == "LNCEL_FDD/NRCELL":
                    if param == "id":
                        parsed["lncel_fdd_id"] = val
                    elif param == "rootSeqIndex":
                        parsed["root_seq_index"] = val
                elif mo == "LNCEL_IOT":
                    if param == "id":
                        parsed["iot_lncel_id"] = val
                    elif param == "lcrId":
                        parsed["iot_lcr_id"] = val
                    elif param == "nbIotLinkedCellId":
                        parsed["iot_nbiot_linked_cell_id"] = val
                    elif param == "tac":
                        parsed["iot_tac"] = val
                elif mo == "Radio" and param == "Type":
                    parsed["radio_type"] = val
                elif mo == "Radio" and param == "Version":
                    parsed["radio_ver"] = val
                    
        return parsed
    
    def _build_script_cell(self, parsed, du_type, prod_map, du_dummy_killed):
        rt = (parsed.get("radio_type") or "").strip().upper()
        if rt == "5G":
            return self._build_script_cell_5g(parsed, du_type, prod_map, du_dummy_killed)
        return self._build_script_cell_lte(parsed, du_type, prod_map, du_dummy_killed)

    
    def _build_script_cell_lte(self, parsed, du_type, prod_map, du_dummy_killed):
        mrbts_id = parsed["mrbts_id"]
        mrbts_ip = parsed["mrbts_ip"]
        mrbts_id_str = str(mrbts_id)
        rmod_id = parsed["rmod_id"]
        prod_name = parsed["prod_name"]
        sector_3 = str(parsed.get("sector_3", "")).strip().upper()
        cablink_id = parsed.get("cablink_id")
        dest_mod = parsed.get("dest_mod")
        first_label = parsed.get("first_endpoint_label")
        first_port = parsed.get("first_endpoint_port_id")
        second_label = parsed.get("second_endpoint_label")
        second_port = parsed.get("second_endpoint_port_id")

        lncel_id = parsed["lncel_id"]
        cell_name = parsed.get("cell_name")
        lcr_id = parsed["lcr_id"]
        phy_cell_id = parsed["phy_cell_id"]
        tac = parsed.get("tac")
        nbiot_linked_cell_id = parsed.get("nbiot_linked_cell_id")
        lncel_fdd_id = parsed.get("lncel_fdd_id")
        root_seq_index = parsed.get("root_seq_index")

        iot_lncel_id = parsed.get("iot_lncel_id")
        iot_lcr_id = parsed.get("iot_lcr_id")
        iot_nbiot_linked_cell_id = parsed.get("iot_nbiot_linked_cell_id")
        iot_tac = parsed.get("iot_tac")

        radio_ver = parsed["radio_ver"]
        radio_type = parsed["radio_type"]
        du_type = parsed.get("du_type", "FSMF")

        script = []
        script += ["### HEADER ###"]
        script += [
            f"set-bts {mrbts_id} {mrbts_ip}",
            f"dest-bts {mrbts_id}",
            f"check-ping {mrbts_id}",
            f"set-mode cell",
            f"set-mo-version {radio_ver}",
            f"set-rat-type {radio_type}",
            f"set-iot-lncel-id {iot_lncel_id}",
            f"set-du-type {du_type}",
            "set-allow-commit-diff true",
            ""
        ]
        if prod_name.strip().upper() == "FXCA" and sector_3 == "TRUE":
            script += ["set-ru-para sector_3 true"]

        script += [
            "### BODY ###",
            f"set-ru-type {prod_name}",
            f"tgt-bts {mrbts_id_str}",
            "dnload-bts-cfg",
            f"set-ru-type {prod_name}",
            "EQM 1", "APEQM 1", f"RMOD {rmod_id}", "auto-config ANTL *",
            f"moduleLocation {cell_name}", "exit-all"
        ]

        if dest_mod.startswith("FHS"):
            script += [
                "EQM 1",
                "HWTOP 1",
                f"CABLINK {cablink_id}",
                f"firstEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/{dest_mod}",
                f"firstEndpointLabel {first_label}",
                f"firstEndpointPortId {first_port}",
                f"secondEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/RMOD-{rmod_id}",
                f"secondEndpointLabel {second_label}",
                f"secondEndpointPortId {second_port}",
                "exit-all"
            ]
        else :
            script += [
                "EQM 1",
                "HWTOP 1",
                f"CABLINK {cablink_id}",
                f"firstEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/CABINET-1/{dest_mod}",
                f"firstEndpointLabel {first_label}",
                f"firstEndpointPortId {first_port}",
                f"secondEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/RMOD-{rmod_id}",
                f"secondEndpointLabel {second_label}",
                f"secondEndpointPortId {second_port}",
                "exit-all"
            ]

        script += [
            f"LNBTS {mrbts_id_str}",
            f"LNCEL {lncel_id}",
            f"lcrId {lcr_id}",
            f"phyCellId {phy_cell_id}",
            f"tac {tac}"
        ]
        if iot_lncel_id :
            script += [
                f"nbIotLinkedCellId {nbiot_linked_cell_id}",
            ]
        else :
            script += [
                "nbIoTMode disabled"
            ]
        script += [
            "LNCEL_FDD 0",
            f"rootSeqIndex {root_seq_index}",
            "exit",
            "auto-config",
            "exit-all",
        ]

        if iot_lncel_id and du_type == "FSMF":
            script += [
                f"LNBTS {mrbts_id_str}",
                f"LNCEL {iot_lncel_id}",
                f"lcrId {iot_lcr_id}",
                f"phyCellId {phy_cell_id}",
                f"tac {iot_tac}",
                f"nbIotLinkedCellId {iot_nbiot_linked_cell_id}",
                "NBIOT_FDD 0",
                "dlChBw '0.2 MHz'", 
                "exit",
                "REDRT 0",
                "exit",
                "DRX 0",
                "exit",
                "SDRX 0",
                "exit",
                "SIB 0",
                "exit-all"
            ]
        elif iot_lncel_id and du_type == "du20":
            script += [
                f"LNBTS {mrbts_id_str}",
                f"LNCEL {iot_lncel_id}",
                f"lcrId {iot_lcr_id}",
                f"phyCellId {phy_cell_id}",
                f"tac {tac}",
                f"nbIotLinkedCellId {iot_nbiot_linked_cell_id}",
                "NBIOT_FDD 0",
                "dlChBw '0.2 MHz'", 
                "exit",
                "BBPOOLALLOC 0",
                "exit",
                "REDRT 0",
                "exit",
                "DRX 0",
                "exit",
                "SDRX 0",
                "exit",
                "SIB 0",
                "exit-all"
            ]

        script += [
            "MNL 1",
            "MNLENT 1",
            "CELLMAPPING 1",
            f"auto-config LCELL {lcr_id}",
        ]
        if iot_lncel_id :
            script += [f"auto-config LCELL {iot_lcr_id}"]
        script += ["exit-all", ""]

        # 더미 제거
        if self.kill_dummy_flag and mrbts_id_str not in du_dummy_killed:
            script += ["# 더미 제거 시작",
                    "EQM 1", "APEQM 1", "no-mo-scf RMOD 32767", "exit-all",
                    "EQM 1", "HWTOP 1", "no-mo-scf CABLINK 65535", "exit-all",
                    f"LNBTS {mrbts_id_str}",
                    "no-mo-scf LNCEL 65535", "no-mo-scf LNCEL 65534", "exit-all",
                    "MNL 1", "MNLENT 1", "CELLMAPPING 1",
                    "no-mo-scf LCELL 254", "no-mo-scf LCELL 255", "no-mo-scf LNMME 10", "no-mo-scf LNMME 11",
                    "exit-all", "# 더미 제거 끝", ""]
            du_dummy_killed.add(mrbts_id_str)

        script += ["commit", "apply-bts-cfg", "#act-bts-cfg"]
        return script
    
    def _build_script_cell_5g(self, parsed, du_type, prod_map, du_dummy_killed):
        mrbts_id = parsed["mrbts_id"]
        mrbts_ip = parsed["mrbts_ip"]
        mrbts_id_str = str(mrbts_id)
        rmod_id = parsed["rmod_id"]
        prod_name = parsed["prod_name"]
        cell_type = parsed["cell_type"]
        cablink_id = parsed.get("cablink_id")
        dest_mod = parsed.get("dest_mod")
        first_label = parsed.get("first_endpoint_label")
        first_port = parsed.get("first_endpoint_port_id")
        second_label = parsed.get("second_endpoint_label")
        second_port = parsed.get("second_endpoint_port_id")

        lncel_id = parsed["lncel_id"]
        cell_name = parsed.get("cell_name")
        lcr_id = parsed["lcr_id"]
        phy_cell_id = parsed["phy_cell_id"]
        tac = parsed.get("tac")
        root_seq_index = parsed.get("root_seq_index")

        iot_lncel_id = parsed.get("iot_lncel_id")

        radio_ver = parsed["radio_ver"]
        radio_type = parsed["radio_type"]
        du_type = parsed.get("du_type", "du10")

        script = []
        script += ["### HEADER ###"]
        script += [
            f"set-bts {mrbts_id} {mrbts_ip}",
            f"dest-bts {mrbts_id}",
            f"check-ping {mrbts_id}",
            f"set-mode cell",
            f"set-mo-version {radio_ver}",
            f"set-rat-type {radio_type}",
            f"set-iot-lncel-id {iot_lncel_id}",
            f"set-du-type {du_type}",
            "set-allow-commit-diff true",
            ""
        ]

        script += [
            "### BODY ###",
            f"set-ru-type {prod_name}",
            f"tgt-bts {mrbts_id_str}",
            "dnload-bts-cfg",
            f"set-ru-type {prod_name}",
            f"set-cell-type {cell_type}",
            "EQM 1", "APEQM 1"
        ]
            
        if prod_name == "APHA":
            script.append(f"ASIRMOD {rmod_id}")
        else:
            script.append(f"RMOD {rmod_id}")

        script.append("exit-all")

        if prod_name == "APHA":
            script += [
                "EQM 1",
                "HWTOP 1",
                f"CABLINK {cablink_id}",
                f"firstEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/CABINET-1/{dest_mod}",
                f"firstEndpointLabel {first_label}",
                f"firstEndpointPortId {first_port}",
                "iqCompression none",
                "linkSpeed Cpri7",
                "radioProtocolType TDDCPRI",
                f"secondEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/ASIRMOD-{rmod_id}",
                f"secondEndpointLabel {second_label}",
                f"secondEndpointPortId {second_port}",
                "exit-all"
            ]
        else :
            script += [
                "EQM 1",
                "HWTOP 1",
                f"CABLINK {cablink_id}",
                f"firstEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/CABINET-1/{dest_mod}",
                f"firstEndpointLabel {first_label}",
                f"firstEndpointPortId {first_port}",
                "linkSpeed eCpri10",
                "radioProtocolType 10eCPRI",
                f"secondEndpointDN MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/RMOD-{rmod_id}",
                f"secondEndpointLabel {second_label}",
                f"secondEndpointPortId {second_port}",
                "exit-all"
            ]

        script += [
            f"NRBTS {mrbts_id_str}",
            f"NRCELL {lncel_id}",
            f"cellName {cell_name}",
            f"lcrId {lcr_id}",
            f"physCellId {phy_cell_id}",
            f"prachRootSequenceIndex {root_seq_index}",
            f"configuredEpsTac {tac}",
            "exit-all", ""
        ]

        script += [
            "MNL 1",
            "MNLENT 1",
            "CELLMAPPING 1",
            f"auto-config LCELNR {lcr_id}",
            "exit-all", ""
        ]

        # 더미 제거
        if self.kill_dummy_flag and mrbts_id_str not in du_dummy_killed:
            script += ["# 더미 제거 시작",
                    "EQM 1", "APEQM 1", "no-mo-scf RMOD 31", "exit-all",
                    "MNL 1", "MNLENT 1", "SYNC 2", f'no-list targetDNList "MRBTS-{mrbts_id_str}/EQM-1/APEQM-1/RMOD-31"', "exit-all",
                    f"NRBTS {mrbts_id_str}",
                    "NRDU 1", "no-list refNrCellGroup 0",
                    "no-mo-scf NRCELLGRP 0", "exit",
                    "no-mo-scf NRPGRP 31", "exit-all",
                    "EQM 1", "HWTOP 1", "no-mo-scf CABLINK 65535", "exit-all",
                    f"LNBTS {mrbts_id_str}",
                    "no-mo-scf NRCELL 6144", "no-mo-scf NRCELL 6144", "exit-all",
                    "MNL 1", "MNLENT 1", "CELLMAPPING 1",
                    "no-mo-scf LCELNR 31",
                    "exit-all", "# 더미 제거 끝", ""]
            du_dummy_killed.add(mrbts_id_str)

        script += ["commit", "apply-bts-cfg", "#act-bts-cfg"]
        return script

    def _parse_excel_rows_init_mod(self, ws):
        script_entries = []

        # [1] CLI SCENARIO 행 탐색
        scenario_row_idx = None
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().upper() == "CLI SCENARIO":
                scenario_row_idx = row_idx
                break
        if scenario_row_idx is None:
            self.perror("엑셀 시트에서 'CLI SCENARIO' 셀을 찾을 수 없습니다.")
            return []

        mo_row_idx = scenario_row_idx
        key_row_idx = mo_row_idx + 1
        data_start_row_idx = mo_row_idx + 2
        max_col = ws.max_column

        mo_headers = [ws.cell(row=mo_row_idx, column=c).value for c in range(1, max_col + 1)]
        param_keys = [ws.cell(row=key_row_idx, column=c).value for c in range(1, max_col + 1)]

        for i in range(len(mo_headers)):
            if mo_headers[i] is None and i > 0:
                mo_headers[i] = mo_headers[i - 1]

        du_type = "du10"
        radiover_col = radiotype_col = None
        for col in range(max_col):
            if mo_headers[col] == "Radio" and param_keys[col] == "Type":
                radiotype_col = col
            elif mo_headers[col] == "Radio" and param_keys[col] == "Version":
                radiover_col = col
        if radiover_col is None or radiotype_col is None:
            self.perror("Radio/Type 또는 Radio/Version 열을 찾을 수 없습니다.")
            return []

        mo_columns = {}
        current_mo = None
        for col in range(max_col):
            mo = mo_headers[col]
            param = param_keys[col]
            if mo:
                current_mo = mo
                if current_mo not in mo_columns:
                    mo_columns[current_mo] = []
            if current_mo and param:
                mo_columns[current_mo].append((param, col))

        data_end_row_idx = ws.max_row + 1
        for row_idx in range(data_start_row_idx, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=4).value
            if val and str(val).strip().startswith("#####"):
                data_end_row_idx = row_idx
                break

        try:
            action_col = mo_headers.index("CLI Scenario")
        except ValueError:
            self.perror("'CLI Scenario' 열이 존재하지 않습니다.")
            return []

        for row_idx in range(data_start_row_idx, data_end_row_idx):
            row = [ws.cell(row=row_idx, column=c + 1).value for c in range(max_col)]
            action_val = row[action_col]
            if not action_val or str(action_val).strip().upper() != "INIT-MOD":
                continue

            mrbts_id = None
            bbmod_list = []
            bbmod_id_temp = bbmod_prod_temp = None
            fhs_id = None
            cablink_l1_id = l1_dest_mod = l1_first_end_point_label = l1_first_end_point_port_id = l1_second_end_point_label = l1_second_end_point_port_id = None
            cablink_l2_id = l2_dest_mod = l2_first_end_point_label = l2_first_end_point_port_id = l2_second_end_point_label = l2_second_end_point_port_id = None
            for mo, key_list in mo_columns.items():
                for param, col in key_list:
                    val = row[col]
                    if mo == "MRBTS" and param == "id":
                        mrbts_id = val
                    elif mo == "BBMOD":
                        if param == "id":
                            bbmod_id_temp = val
                        elif param == "prod-name":
                            bbmod_prod_temp = val
                        if bbmod_id_temp and bbmod_prod_temp:
                            bbmod_list.append((bbmod_id_temp, bbmod_prod_temp))
                            bbmod_id_temp = bbmod_prod_temp = None
                    elif mo == "FHS_1" and param == "id":
                        fhs_id = val
                    elif mo == "CABLINK_L1":
                        if param == "id":
                            cablink_l1_id = val
                        elif param == "dest-mod":
                            l1_dest_mod = val
                        elif param == "firstEndpointLabel":
                            l1_first_end_point_label = val
                        elif param == "firstEndpointPortId":
                            l1_first_end_point_port_id = val
                        elif param == "secondEndpointLabel":
                            l1_second_end_point_label = val
                        elif param == "secondEndpointPortId":
                            l1_second_end_point_port_id = val
                    elif mo == "CABLINK_L2":
                        if param == "id":
                            cablink_l2_id = val
                        elif param == "dest-mod":
                            l2_dest_mod = val
                        elif param == "firstEndpointLabel":
                            l2_first_end_point_label = val
                        elif param == "firstEndpointPortId":
                            l2_first_end_point_port_id = val
                        elif param == "secondEndpointLabel":
                            l2_second_end_point_label = val
                        elif param == "secondEndpointPortId":
                            l2_second_end_point_port_id = val

            radio_type = row[radiotype_col]
            radio_ver = row[radiover_col]
            if not (mrbts_id):
                continue

            mrbts_id_str = str(mrbts_id)

            script = []
            script += ["### HEADER ###"]
            script += ["exit", "yes"]
            script += [
                f"dest-bts {mrbts_id}",
                f"check-ping {mrbts_id}",
                f"set-mode mod",
                f"set-mo-version {radio_ver}",
                f"set-rat-type {radio_type}",
                "set-allow-commit-diff true",  # 필요 시 설정
                ""
            ]

            script += [
                "### BODY ###",
                f"set-du-type {du_type}",
                f"set-ru-type AHCA", ## 수정 필요
                f"tgt-bts {mrbts_id_str}",
                "dnload-bts-cfg",
                "exit-all"
            ]
            if bbmod_list:
                script += ["# BBMOD 설정", "EQM 1", "APEQM 1"]
                for bb_id, bb_prod in bbmod_list:
                    script += [f"BBMOD {bb_id}", f"prodCodePlanned {bb_prod}"]
                script += ["exit-all", "# BBMOD 설정 종료", ""]

            script.append("")

            script_entries.append((mrbts_id_str, script))

        return script_entries

    def _write_script_files(self, bts_entries, cell_entries, mod_entries):
        today_str = datetime.now().strftime("%Y%m%d")

        # autocomm/today_str 경로 생성
        output_dir = get_path(self.env_type, "autocomm", today_str)
        os.makedirs(output_dir, exist_ok=True)

        all_errors = []

        # Step 1: DU 시퀀스 초기화
        du_seq_map = OrderedDict()
        seq_counter = 1

        # Step 2: BTS 스크립트 저장 + DU 시퀀스 부여
        for du_id, script_lines, _, fail, error_lines in bts_entries:
            if du_id not in du_seq_map:
                du_seq_map[du_id] = seq_counter * 100
                seq_counter += 1

            seq_num = du_seq_map[du_id]
            postfix = "fail" if fail else "pass"
            filename = f"{seq_num}.DU{du_id}_x_initbts_{postfix}.cli"
            filepath = os.path.join(output_dir, filename)

            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(script_lines) + "\n")
            self.poutput(f"[{filename}] 생성 완료")
            save_to_server(self, output_path=filepath, purpose="autocomm")

            for line in error_lines:
                all_errors.append(f"[{filename}] {line}")

        # Step 3: Cell 스크립트 저장
        cell_index_map = {du_id: 1 for du_id in du_seq_map}
        for du_id, script_lines, filename_type, fail, error_lines in cell_entries:
            if du_id not in du_seq_map:
                du_seq_map[du_id] = seq_counter * 100
                cell_index_map[du_id] = 1
                seq_counter += 1

            cell_id = filename_type.split("-")[1]
            seq_num = du_seq_map[du_id] + cell_index_map[du_id]
            postfix = "fail" if fail else "pass"
            filename = f"{seq_num}.DU{du_id}_Cell{cell_id}_initcell_{postfix}.cli"
            cell_index_map[du_id] += 1

            filepath = os.path.join(output_dir, filename)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(script_lines) + "\n")
            self.poutput(f"[{filename}] 생성 완료")
            save_to_server(self, output_path=filepath, purpose="autocomm")

            for line in error_lines:
                all_errors.append(f"[{filename}] {line}")

        # Step 4: Mod 스크립트 저장
        mod_index_map = {du_id: cell_index_map.get(du_id, 1) for du_id in du_seq_map}
        for du_id, script_lines in mod_entries:
            if du_id not in du_seq_map:
                du_seq_map[du_id] = seq_counter * 100
                mod_index_map[du_id] = 1
                seq_counter += 1

            seq_num = du_seq_map[du_id] + mod_index_map[du_id]
            filename = f"{seq_num}.DU{du_id}_initmod_pass.cli"
            mod_index_map[du_id] += 1

            filepath = os.path.join(output_dir, filename)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(script_lines) + "\n")
            self.poutput(f"[{filename}] 생성 완료")
            save_to_server(self, output_path=filepath, purpose="autocomm")

        # Step 5: 에러 로그 저장 + 서버 전송
        if all_errors:
            log_path = os.path.join(output_dir, "parse_errors.log")
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(log_path, "a", encoding="utf-8") as f:
                for line in all_errors:
                    f.write(f"[{now_str}] {line}\n")
            self.poutput(f"[parse_errors.log] {len(all_errors)}건 저장됨")

            # 서버에 업로드
            save_to_server(self, output_path=log_path, purpose="autocomm")


    def do_init_sw_ver(self, arg):
        """
        여러 BTS에 대해 소프트웨어를 병렬 업데이트합니다.
        사용법: init-sw-ver <엑셀파일명>
        """
        excel_filename = arg.strip()
        if not excel_filename:
            self.perror("엑셀 파일명을 입력하세요.")
            return

        try:
            excel_bytes = load_from_server(filename=excel_filename, filetype="binary", purpose="autocomm")
            if not excel_bytes:
                self.perror(f"서버에서 엑셀 파일을 불러오지 못했습니다: {excel_filename}")
                return

            wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
            ws_bts = None
            for sheet_name in wb.sheetnames:
                if "INITBTS" in sheet_name.upper():
                    ws_bts = wb[sheet_name]
                    break
            if ws_bts is None:
                raise KeyError("INITBTS 시트를 찾을 수 없습니다.")
            entries = self._extract_bts_sw_entries(ws_bts)
        except Exception as e:
            self.perror(f"엑셀 처리 오류: {e}")
            return

        if not entries:
            self.poutput("업데이트 대상 BTS가 없습니다.")
            return

        payload = json.dumps(entries)
        request = message_pb2.Request(command="init-sw-ver", payload=payload)

        try:
            response = grpc_stub.SendCommand(request)
            if response.success:
                try:
                    result_list = json.loads(response.result)
                    success, fail, skipped = 0, 0, 0
                    failed_entries = []

                    for item in result_list:
                        status = item.get("status")
                        msg = item.get("message", "")
                        bts_id = item.get("bts_id", "unknown")
                        if status == "success":
                            success += 1
                            self.poutput(f"[{bts_id}] 성공: {msg}")
                        elif status == "fail":
                            fail += 1
                            failed_entries.append((bts_id, msg))
                            self.perror(f"[{bts_id}] 실패: {msg}")
                            
                            raw = item.get("raw")
                            if raw:
                                self.poutput(f"[{bts_id}] 🔍 상세 로그:\n{raw}")
                        elif status == "skipped":
                            skipped += 1
                            self.poutput(f"[{bts_id}] 스킵됨: {msg}")

                    self.poutput(f"\n총 {len(result_list)}건 중: 성공 {success}건 / 실패 {fail}건 / 스킵 {skipped}건")

                    if failed_entries:
                        self.poutput("\n 실패 목록 요약:")
                        for bts_id, msg in failed_entries:
                            self.poutput(f" - [{bts_id}]: {msg}")

                except json.JSONDecodeError:
                    self.poutput("결과 파싱 실패 (원시 결과 출력):")
                    self.poutput(response.result)
            else:
                self.perror(f"실패: {response.result}")
        except Exception as e:
            self.perror(f"gRPC 오류: {e}")


    def _extract_bts_sw_entries(self, ws_bts):
        """
        INITBTS 시트에서 MRBTS id와 Radio Version 값을 추출합니다.
        """
        mo_row_idx = key_row_idx = None

        # 헤더 행 탐색
        for row_idx in range(1, ws_bts.max_row + 1):
            val = ws_bts.cell(row=row_idx, column=4).value
            if val and str(val).strip().upper() == "CLI SCENARIO":
                mo_row_idx = row_idx
                key_row_idx = mo_row_idx + 1
                break
        if mo_row_idx is None:
            raise ValueError("엑셀 시트에서 'CLI SCENARIO' 셀을 찾을 수 없습니다.")

        data_start_row_idx = mo_row_idx + 2
        max_col = ws_bts.max_column

        # 데이터 종료 행 찾기 (##### 주석줄 기준)
        data_end_row_idx = ws_bts.max_row
        for row_idx in range(data_start_row_idx, ws_bts.max_row + 1):
            val = ws_bts.cell(row=row_idx, column=4).value
            if val and str(val).strip().startswith("#####"):
                data_end_row_idx = row_idx - 1
                break

        # 헤더 구성
        mo_headers = [ws_bts.cell(row=mo_row_idx, column=c).value for c in range(1, max_col + 1)]
        param_keys = [ws_bts.cell(row=key_row_idx, column=c).value for c in range(1, max_col + 1)]

        # 빈 MO 채우기
        for i in range(len(mo_headers)):
            if mo_headers[i] is None and i > 0:
                mo_headers[i] = mo_headers[i - 1]

        # 대상 열 찾기
        mrbts_col = version_col = None
        for col in range(max_col):
            if mo_headers[col] == "MRBTS" and param_keys[col] == "id":
                mrbts_col = col
            elif mo_headers[col] == "Radio" and param_keys[col] == "Version":
                version_col = col
        if mrbts_col is None or version_col is None:
            raise ValueError("MRBTS id 또는 Radio Version 열을 찾을 수 없습니다.")

        # 데이터 추출 (지정된 범위 내에서만)
        entries = []
        for row_idx in range(data_start_row_idx, data_end_row_idx + 1):
            cli_scenario = ws_bts.cell(row=row_idx, column=4).value
            if str(cli_scenario).strip().upper() != "INIT-BTS":
                continue

            mrbts_id = ws_bts.cell(row=row_idx, column=mrbts_col + 1).value
            sw_ver = ws_bts.cell(row=row_idx, column=version_col + 1).value

            if mrbts_id and sw_ver:
                entries.append((str(mrbts_id).strip(), str(sw_ver).strip()))

        return entries

    def do_show_bts_entry(self, arg):
        """
        각 BTS의 현재 SCF 버전과 엑셀에 정의된 목표 버전을 비교하여 출력합니다.
        사용법: show-bts-entry <엑셀파일명>
        """
        excel_filename = arg.strip()
        if not excel_filename:
            self.perror("사용법: show-bts-entry <엑셀파일명>")
            return

        try:
            excel_bytes = load_from_server(filename=excel_filename, filetype="binary", purpose="autocomm")
            if not excel_bytes:
                self.perror(f"[오류] 서버에서 엑셀 파일을 불러오지 못했습니다: {excel_filename}")
                return
            wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
            
            ws_bts = None
            for sheet_name in wb.sheetnames:
                if "INITBTS" in sheet_name.upper():
                    ws_bts = wb[sheet_name]
                    break
            if ws_bts is None:
                raise KeyError("INITBTS 시트를 찾을 수 없습니다.")
        except Exception as e:
            self.perror(f"[오류] 엑셀 파일 열기 실패: {e}")
            return

        # 헤더 찾기
        mo_row_idx = key_row_idx = None
        for row_idx in range(1, ws_bts.max_row + 1):
            val = ws_bts.cell(row=row_idx, column=4).value
            if val and str(val).strip().upper() == "CLI SCENARIO":
                mo_row_idx = row_idx
                key_row_idx = mo_row_idx + 1
                break
        if mo_row_idx is None:
            self.perror("엑셀에서 'CLI SCENARIO' 헤더를 찾을 수 없습니다.")
            return

        data_start_row_idx = mo_row_idx + 2
        max_col = ws_bts.max_column
        mo_headers = [ws_bts.cell(row=mo_row_idx, column=c).value for c in range(1, max_col + 1)]
        param_keys = [ws_bts.cell(row=key_row_idx, column=c).value for c in range(1, max_col + 1)]

        for i in range(1, len(mo_headers)):
            if mo_headers[i] is None:
                mo_headers[i] = mo_headers[i - 1]

        # ID, Version 열 찾기
        btsid_col = version_col = None
        for col in range(max_col):
            if mo_headers[col] == "MRBTS" and param_keys[col] == "id":
                btsid_col = col
            elif mo_headers[col] == "Radio" and param_keys[col] == "Version":
                version_col = col
        if btsid_col is None or version_col is None:
            self.perror("엑셀에서 MRBTS.id 또는 Radio.Version 열을 찾을 수 없습니다.")
            return

        # 유효 데이터 끝 찾기
        data_end_row_idx = ws_bts.max_row + 1
        for row_idx in range(data_start_row_idx, ws_bts.max_row + 1):
            val = ws_bts.cell(row=row_idx, column=4).value
            if val and str(val).strip().startswith("#####"):
                data_end_row_idx = row_idx
                break

        self.poutput(f"{'BTS ID':<10} | {'현재 버전':<10} | {'목표 버전':<10}")
        self.poutput("-" * 35)

        seen_bts_ids = set()

        for row_idx in range(data_start_row_idx, data_end_row_idx):
            row = [ws_bts.cell(row=row_idx, column=c + 1).value for c in range(max_col)]
            bts_id = row[btsid_col]
            target_version = row[version_col]

            if not bts_id or not target_version:
                continue

            bts_id_str = str(bts_id).strip()
            if not bts_id_str.isdigit():
                continue
            if bts_id_str in seen_bts_ids:
                continue
            seen_bts_ids.add(bts_id_str)

            try:
                self.do_dest_bts(bts_id_str)
                self.do_dnload_bts_cfg_raw("")
                self.do_set_cfg_scf("genScf")
                version = self._extract_version_from_xml()
            except Exception as e:
                self.perror(f"[{bts_id_str}] 처리 중 오류: {e}")
                continue

            self.poutput(f"{bts_id_str:<10} | {version:<10} | {target_version:<10}")
        

    def _extract_version_from_xml(self) -> str:
        """
        현재 self.xml_tree에서 EQM 클래스의 version 값을 추출하여 '25R2' 등의 형태로 반환
        """
        if not self.xml_tree:
            raise ValueError("xml_tree가 존재하지 않습니다.")
        
        root = self.xml_tree.getroot()
        for mo in root.iter("managedObject"):
            if mo.attrib.get("class") == "com.nokia.srbts.eqm:EQM":
                ver = mo.attrib.get("version")
                if ver:
                    match = re.search(r"(\d+R\d+)", ver)  # 예: EQM25R2 → 25R2 추출
                    if match:
                        return match.group(1)
        raise ValueError("EQM 클래스에서 version 정보를 추출할 수 없습니다.")

    def do_copy_ac_tmpl(self, arg):
        """
        사용법: copy-ac-tmpl "<로컬파일명>" "<서버디렉토리 autocomm하위>"
        """
        tokens = shlex.split(arg)
        if len(tokens) != 2:
            self.perror("사용법: copy-ac-tmpl <로컬파일명> <서버디렉토리 autocomm하위>")
            return

        local_path, server_dir = tokens
        if not os.path.exists(local_path):
            self.perror(f"[오류] 로컬 파일이 존재하지 않음: {local_path}")
            return

        file_name = os.path.basename(local_path)
        server_path = os.path.join(server_dir, file_name)  # "temp/CLI 개통탬플릿.xlsx"처럼 조합

        with open(local_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("utf-8")

        payload = f"autocomm||{server_path}||{encoded}"
        req = message_pb2.Request(command="saveFile", payload=payload)
        resp = grpc_stub.SendCommand(req)

        if resp.success:
            self.poutput(f"[copy-ac-tmpl] 업로드 성공: {server_path}")
        else:
            self.perror(f"[copy-ac-tmpl] 실패: {resp.result}")

    def do_del_ac_tmpl(self, arg):
        """
        서버에서 autocomm 템플릿 파일을 삭제합니다.
        사용법: del-ac-tmpl <서버상 상대경로>
        예시: del-ac-tmpl temp/CLI_개통템플릿.xlsx
        """
        tokens = shlex.split(arg)
        if len(tokens) != 1:
            self.perror("사용법: del-ac-tmpl <서버상 상대경로>")
            return

        filename = tokens[0]
        delete_from_server(self, filename=filename, purpose="autocomm")
